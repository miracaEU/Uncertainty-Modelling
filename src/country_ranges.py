"""Per-country, per-scenario EAD uncertainty ranges for one asset type.

Companion to src/eu_totals.py. That module sums countries into one European
figure per scenario; this one keeps the country dimension and reports how WIDE
each national distribution is - the p5-p95 band and the usual central
statistics behind it.

Everything here is presentation over numbers src/ead_ranges.py already wrote to
overview_figures/ead_ranges/EAD_Ranges.csv. No model runs, no archive reads.

THREE WAYS TO SAY "HOW WIDE"
----------------------------
    w90_EUR         p95 - p5, in euro. The band itself. Comparable across
                    scenarios for one country, but a big country always looks
                    more uncertain than a small one.
    factor_p95_p5   p95 / p5. Dimensionless - "the plausible range spans a
                    factor of N". Undefined where p5 is zero.
    w90_rel         (p95 - p5) / median. Also dimensionless, and it survives a
                    zero p5, so it is the one column that is populated for every
                    row. This is the definition src/ead_ranges.py uses.

Outputs (default overview_figures/{asset}/):
    EAD_country_ranges_{asset}.xlsx
        Ranges            one row per (country, scenario) - every statistic
        Spread_matrix     countries x scenarios, factor_p95_p5
        Width_matrix      countries x scenarios, w90_rel
        Median_matrix     countries x scenarios, median EUR/yr
        Scenario_summary  per scenario: spread across countries, widest/narrowest
        Notes             what is and is not safe to read off these numbers

Usage:
    python -m src.country_ranges
    python -m src.country_ranges --asset roads
"""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd

from .ead_ranges import hazard_of
from .eu_totals import RANGES_CSV
from .plot_pyramid import PROJECT_ROOT, scenario_label, scenario_sort_key

EUR_BN = 1e9

# Columns carried straight through from EAD_Ranges.csv, renamed to make the unit
# explicit. Order here is the order in the sheet: the band first, then the shape
# of the distribution, then the reference comparison.
PASSTHROUGH = {
    "p5": "p5_EUR",
    "p25": "p25_EUR",
    "median": "median_EUR",
    "p75": "p75_EUR",
    "p95": "p95_EUR",
    "mean": "mean_EUR",
    "min": "min_EUR",
    "max": "max_EUR",
    "std": "std_EUR",
    "w90": "w90_EUR",
    "w50": "w50_EUR",
}


def _safe_div(num: np.ndarray, den: np.ndarray) -> np.ndarray:
    """num/den, NaN wherever the denominator is not strictly positive.

    np.where evaluates both branches, so the guard has to go in the division
    itself or a zero denominator still raises before the selection happens.
    """
    den = np.asarray(den, dtype=float)
    ok = den > 0
    return np.divide(np.asarray(num, dtype=float), np.where(ok, den, 1.0),
                     out=np.full(den.shape, np.nan), where=ok)


def build_rows(asset: str, ranges_csv: Path = RANGES_CSV) -> pd.DataFrame:
    """One row per (country, scenario) for `asset`, with the spread measures."""
    if not ranges_csv.exists():
        raise SystemExit(f"No {ranges_csv} - run `python -m src.ead_ranges` first.")
    df = pd.read_csv(ranges_csv)
    sub = df[df["asset"] == asset].copy()
    if sub.empty:
        raise SystemExit(f"No rows for asset '{asset}' in {ranges_csv.name}.")

    out = pd.DataFrame({
        "country": sub["country"],
        "scenario": sub["scenario"],
        "scenario_label": sub["scenario"].map(scenario_label),
        "hazard": sub["scenario"].map(hazard_of),
        "n_draws": sub["n"],
        "source": sub["source"],
    })
    for src, dest in PASSTHROUGH.items():
        out[dest] = sub[src]

    med = sub["median"].to_numpy(dtype=float)
    p5 = sub["p5"].to_numpy(dtype=float)
    p95 = sub["p95"].to_numpy(dtype=float)
    mean = sub["mean"].to_numpy(dtype=float)
    std = sub["std"].to_numpy(dtype=float)

    # p5 is exactly zero wherever more than 5% of draws give no damage at all -
    # a country fully protected under most protection draws. The ratio is then
    # undefined rather than infinite: the band has no lower bound to divide by.
    out["factor_p95_p5"] = _safe_div(p95, p5)
    out["w90_rel"] = sub["w90_rel"]
    out["w50_rel"] = _safe_div(sub["p75"].to_numpy(dtype=float) - sub["p25"].to_numpy(dtype=float), med)
    out["cv"] = _safe_div(std, mean)
    out["mean_median_ratio"] = sub["mean_median_ratio"]
    out["zero_fraction"] = sub["zero_fraction"]
    out["p5_is_zero"] = sub["p5_is_zero"]

    out["ref_mid_EUR"] = sub["ref_mid"]
    out["ref_percentile"] = sub["ref_percentile"]
    out["ref_in_p5_p95"] = sub["ref_in_p5_p95"]
    out["ref_over_median"] = sub["ref_over_median"]

    out = out.sort_values(
        ["scenario", "median_EUR"],
        key=lambda s: s.map(scenario_sort_key) if s.name == "scenario" else s,
        ascending=[True, False],
    )
    return out.reset_index(drop=True)


def country_order(rows: pd.DataFrame) -> list[str]:
    """Countries by total median EAD across scenarios, largest first.

    Puts the countries that carry the European total at the top of every matrix,
    which is how these are read in practice. Re-sorting in Excel is one click if
    an alphabetical view is wanted instead.
    """
    tot = rows.groupby("country")["median_EUR"].sum().sort_values(ascending=False)
    return list(tot.index)


def matrix(rows: pd.DataFrame, value: str) -> pd.DataFrame:
    """Countries (rows) x scenarios (columns) for one statistic."""
    m = rows.pivot(index="country", columns="scenario", values=value)
    cols = sorted(m.columns, key=scenario_sort_key)
    m = m.reindex(index=country_order(rows), columns=cols)
    m.columns = [scenario_label(c, short=True) for c in cols]
    return m.reset_index()


def scenario_summary(rows: pd.DataFrame) -> pd.DataFrame:
    """Per scenario: how the width varies across countries, and the extremes."""
    recs = []
    for scen, g in rows.groupby("scenario"):
        w = g["w90_rel"].dropna()
        f = g["factor_p95_p5"].dropna()
        widest = g.loc[w.idxmax(), "country"] if not w.empty else None
        narrowest = g.loc[w.idxmin(), "country"] if not w.empty else None
        # w90_rel divides by the median, so a country with almost no exposure can
        # post an enormous width off a near-zero denominator (Andorra reaches 390
        # in earthquake). Carrying the widest country's own median lets the reader
        # see at once whether the maximum is a real finding or a small-denominator
        # artefact; w90_rel_median is the robust summary.
        widest_med = float(g.loc[w.idxmax(), "median_EUR"]) if not w.empty else np.nan
        recs.append({
            "scenario": scen,
            "scenario_label": scenario_label(scen),
            "hazard": hazard_of(scen),
            "n_countries": g["country"].nunique(),
            "n_draws": int(g["n_draws"].max()),
            # The EU row: comonotonic bracket, same caveat as eu_totals.py.
            "EU_sum_p5_bn": g["p5_EUR"].sum() / EUR_BN,
            "EU_sum_median_bn": g["median_EUR"].sum() / EUR_BN,
            "EU_sum_p95_bn": g["p95_EUR"].sum() / EUR_BN,
            "EU_sum_mean_bn": g["mean_EUR"].sum() / EUR_BN,
            "w90_rel_min": w.min() if not w.empty else np.nan,
            "w90_rel_median": w.median() if not w.empty else np.nan,
            "w90_rel_max": w.max() if not w.empty else np.nan,
            "factor_p95_p5_median": f.median() if not f.empty else np.nan,
            "n_countries_p5_zero": int(g["p5_is_zero"].sum()),
            "widest_country": widest,
            "widest_country_median_EUR": widest_med,
            "narrowest_country": narrowest,
        })
    out = pd.DataFrame(recs)
    return out.sort_values("scenario", key=lambda s: s.map(scenario_sort_key)).reset_index(drop=True)


NOTES = [
    ("What this file is",
     "One row per country and scenario for a single asset type, showing how wide the "
     "plausible EAD range is. Source: overview_figures/ead_ranges/EAD_Ranges.csv, which "
     "src/ead_ranges.py writes from the Sobol A+B sample."),
    ("Units", "Columns ending _EUR are euro per year. Columns ending _bn are euro billion "
              "per year. Ratio columns are dimensionless."),
    ("w90_EUR", "p95 minus p5, the width of the central 90% band in euro."),
    ("factor_p95_p5",
     "p95 divided by p5 - 'the plausible range spans a factor of N'. Blank where p5 is "
     "zero, which happens when more than 5% of draws give no damage at all (a country "
     "protected out of the hazard under most protection draws). The band then has no "
     "lower bound to divide by; use w90_rel for those rows."),
    ("w90_rel",
     "(p95 - p5) / median. The width measure that survives a zero p5, so it is populated "
     "for every row where the median itself is positive. This is the definition used in "
     "the EAD range figures."),
    ("cv", "std / mean. Sensitive to the extreme right tail, so it runs much larger than "
           "w90_rel; the two are not interchangeable."),
    ("Small countries inflate w90_rel",
     "w90_rel divides by the median, so a country with almost no exposed asset can post a "
     "huge width off a near-zero denominator - Andorra reaches 390 under earthquake on a "
     "median of a few hundred euro. Scenario_summary carries the widest country's own "
     "median so this is visible; w90_rel_median across countries is the robust headline "
     "and sits at 5-13 depending on scenario."),
    ("Skew", "These distributions are strongly right-skewed - mean_median_ratio is "
             "typically 1.5-2. The mean is not the typical outcome."),
    ("What the range is and is not",
     "It is EPISTEMIC: the spread produced by not knowing the vulnerability curve, unit "
     "costs, protection standard, hazard intensity and climate future. It is NOT "
     "interannual variability - EAD is already an expectation over event frequency."),
    ("Ranges are per country, not additive",
     "Do not add p5 or p95 down a column and call it a European interval. Summing "
     "percentiles assumes every country sits at the same percentile at once. The EU_sum_* "
     "columns in Scenario_summary do exactly that and are labelled a comonotonic bracket; "
     "see src/eu_totals.py, and src/cascade.py for the correlation-correct route."),
    ("Scenario naming",
     "'protection as recorded' holds each country at its FLOPROS / COASTPROS design "
     "standard; 'protection sampled' draws the standard as an uncertain factor. Neither "
     "sets protection to zero."),
    ("Reference columns",
     "ref_mid_EUR is the MIRACA_RISK deterministic run for that country and hazard "
     "(climate=current). ref_percentile is where it falls in our sample; ref_in_p5_p95 "
     "says whether it lands inside the central 90% band."),
    ("KNOWN BIAS in the absolute euro columns",
     "Polygon exposure is inflated by 1/cos(latitude) - about 1.29x in Greece to 2.19x in "
     "Sweden - through a cell-area bug in damagescanner (vector.py _get_cell_area_m2). "
     "All _EUR and _bn columns for polygon-bearing assets are affected and will change on "
     "the rerun."),
    ("What the bias does NOT change",
     "The bias is a per-country multiplicative constant, and for power 96.5-99.3% of "
     "flood, coastal and earthquake damage sits on polygons while windstorm is 0% polygon. "
     "A constant cancels in every ratio column here - factor_p95_p5, w90_rel, w50_rel, cv, "
     "mean_median_ratio - so the RELATIVE width results stand. Absolute euro values do "
     "not."),
]


def write_excel(path: Path, rows: pd.DataFrame, summary: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    sheets = {
        "Ranges": rows,
        "Spread_matrix": matrix(rows, "factor_p95_p5"),
        "Width_matrix": matrix(rows, "w90_rel"),
        "Median_matrix": matrix(rows, "median_EUR"),
        "Scenario_summary": summary,
        "Notes": pd.DataFrame(NOTES, columns=["Topic", "Note"]),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        for name, frame in sheets.items():
            frame.to_excel(xl, sheet_name=name, index=False)
            _format_sheet(xl.sheets[name], frame, name)


def _format_sheet(ws, frame: pd.DataFrame, name: str) -> None:
    """Freeze the header, size the columns, and give money and ratios formats."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "B2" if name != "Notes" else "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=(name == "Notes"))

    for i, col in enumerate(frame.columns, start=1):
        letter = get_column_letter(i)
        col_s = str(col)
        if name == "Notes":
            ws.column_dimensions[letter].width = 26 if i == 1 else 110
            for cell in ws[letter][1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            continue

        if col_s.endswith("_EUR") or (name == "Median_matrix" and i > 1):
            fmt, width = "#,##0", 16
        elif col_s.endswith("_bn"):
            fmt, width = "0.000", 15
        elif col_s.startswith(("factor_", "w90_rel", "w50_rel", "cv", "mean_median",
                               "zero_fraction", "ref_percentile", "ref_over_median")):
            fmt, width = "0.00", 15
        elif name in ("Spread_matrix", "Width_matrix") and i > 1:
            fmt, width = "0.00", 15
        elif col_s in ("scenario_label", "widest_country", "narrowest_country"):
            fmt, width = None, 34
        else:
            fmt, width = None, max(10, min(22, len(col_s) + 3))
        ws.column_dimensions[letter].width = width
        if fmt:
            for cell in ws[letter][1:]:
                cell.number_format = fmt


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--asset", default="power", help="asset type (default: power)")
    parser.add_argument("--scenarios", nargs="+", default=None,
                        help="restrict to these scenarios (default: all present)")
    parser.add_argument("--out-dir", default=None, help="default: overview_figures/{asset}")
    args = parser.parse_args()

    rows = build_rows(args.asset)
    if args.scenarios:
        rows = rows[rows["scenario"].isin(args.scenarios)].reset_index(drop=True)
        if rows.empty:
            raise SystemExit(f"No rows left after --scenarios {args.scenarios}.")
    summary = scenario_summary(rows)

    out_dir = Path(args.out_dir) if args.out_dir else PROJECT_ROOT / "overview_figures" / args.asset
    xlsx = out_dir / f"EAD_country_ranges_{args.asset}.xlsx"
    write_excel(xlsx, rows, summary)

    show = summary[["scenario", "n_countries", "EU_sum_p5_bn", "EU_sum_median_bn",
                    "EU_sum_p95_bn", "w90_rel_min", "w90_rel_median", "w90_rel_max",
                    "n_countries_p5_zero", "widest_country"]]
    print(f"{args.asset}: {len(rows)} (country, scenario) rows, "
          f"{rows['country'].nunique()} countries, {rows['scenario'].nunique()} scenarios")
    print(show.to_string(index=False, float_format=lambda v: f"{v:8.2f}"))
    print(f"\nwrote {xlsx}")


if __name__ == "__main__":
    main()