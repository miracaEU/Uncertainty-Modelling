"""Aggregate every Sobol/feature-score result across the whole study into one
Excel workbook, with a legend sheet explaining every abbreviation.

Scans results/ for every {country}_{asset}_{scenario}_sobol_indices.csv and
_feature_scores.csv it can parse (so it picks up whatever combinations have
actually been run - safe to call after a partial study, and automatically
includes new countries/assets/scenarios added later with no code changes).

Produces MIRACA_uncertainty_study_summary.xlsx in the project root with:
  Legend              - what every column, scenario, asset, and factor means
  Curve_Groups        - which curve IDs and object types sit behind every
                        curve factor, per asset/hazard (the source of truth
                        the readable factor names elsewhere are derived from)
  Top_Drivers         - one row per (country, asset, scenario): #1/#2/#3
                        drivers of total_EAD_MEUR by Sobol total effect
  ST_<scenario>       - one heatmap sheet per scenario: factors x
                        (asset, country), Sobol ST on total_EAD_MEUR,
                        colour-scaled, with a border around each column's
                        single most decisive (highest-ST) cell so it's
                        immediately visible whether the same factor wins
                        across every experiment in a scenario or whether it
                        differs
  All_Sobol_Indices   - the full long-format table (every factor, every
                        outcome, every combination) - the raw material
                        everything else here is derived from; pivot this
                        yourself in Excel for anything not already covered
  All_Feature_Scores  - the same, for the LHS extra-trees importance scores
                        (a quicker, less rigorous complement to Sobol)
  Sobol_Convergence   - one row per adaptive-Sobol search: the N sequence it
                        stepped through, where it stopped and why (from
                        results/sobol_convergence_log.jsonl)
  Reruns_Overview     - only present once some combination has been re-run
                        at a higher Sobol sample size N: one row per combo,
                        old vs new #1 driver and its ST - "did more runs
                        change the headline answer"
  Reruns_Detail       - same reruns, but every factor (not just the top
                        one): old vs new S1/ST/ST_conf and the confidence-
                        to-estimate ratio, flagging anything still imprecise
                        even at the higher N
  Timing_By_Combo     - one row per (country, asset): preprocessing/validation
                        time plus the summed total across every scenario -
                        "how long did this country+asset take overall"
  Timing_Summary      - one row per (country, asset, scenario): time spent in
                        each of the LHS/Sobol run+analyze steps, plus a total -
                        "how long did each experiment take"
  Timing_Raw          - the untouched run_study.py log (results/run_study_log.jsonl),
                        one row per logged step, for exact history/troubleshooting

Timing sheets are only written if results/run_study_log.jsonl exists (i.e.
run_study.py has been run at least once) - a manual one-off
run_experiments.py call outside the orchestrator isn't logged there.

Curve factor names (e.g. "curve_F7_4") are replaced everywhere with a
readable description (e.g. "Flood: motorway_trunk/primary (F7.4, 4 curves)")
derived live from src.curves.ASSET_CONFIGS - the same curve group can mean a
different thing in a different asset (curve_F9_1 is "aerodrome/apron" for
airports but "terminal" for power), so the description always uses the
row's own asset for context; heatmap rows (which have no per-cell asset
column) are prefixed with the asset name instead.

This is a derived summary, safe to regenerate at any time - it never writes
to the underlying experiment archives or per-combination CSVs, only reads
them. Every sheet except Reruns_* comes purely from the small CSVs
analyze.py/analyze_sobol.py already produce; the Reruns_* sheets additionally
re-run SALib's Sobol analysis directly on whichever experiments_*.tar.gz
archives are found (see find_sobol_rerun_pairs), since the per-combo indices
CSV is overwritten in place on every rerun and keeps no history of its own.
run_study.py regenerates the whole workbook automatically after a study run.

Usage:
    python -m src.aggregate_results
    python -m src.aggregate_results --output my_summary.xlsx
"""

import argparse
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .curves import ASSET_CONFIGS
from .ema_model import SCENARIOS, build_model
from .paths import PROJECT_ROOT, load_config, set_asset_override, set_country_override, set_scenario_override

OUTPUT_NAME = "MIRACA_uncertainty_study_summary.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1C5CAB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="0B0B0B")
SUBTITLE_FONT = Font(bold=True, size=11, color="256ABF")
WRAP = Alignment(wrap_text=True, vertical="top")
WHITE_FONT = Font(color="FFFFFF")
BLACK_FONT = Font(color="0B0B0B")
TOP_FACTOR_BORDER = Border(*(Side(style="thick", color="0B0B0B") for _ in range(4)))

# Fixed 0/0.5/1 domain (not percentile-relative) so colour intensity is
# comparable across every column/sheet, and so a single text-colour
# threshold (see WHITE_TEXT_THRESHOLD) reliably predicts a dark cell.
FRACTION_COLOR_SCALE = ColorScaleRule(
    start_type="num", start_value=0, start_color="FCFCFB",
    mid_type="num", mid_value=0.5, mid_color="86B6EF",
    end_type="num", end_value=1, end_color="0D366B",
)
# Relative scale for columns that aren't a bounded [0,1] fraction
# (confidence half-widths, interaction = ST-S1 which can be negative).
RELATIVE_COLOR_SCALE = ColorScaleRule(
    start_type="min", start_color="FCFCFB",
    mid_type="percentile", mid_value=50, mid_color="86B6EF",
    end_type="max", end_color="0D366B",
)

# Both colour scales above interpolate through the SAME three stops
# (FCFCFB -> 86B6EF -> 0D366B); only what min/mid/max map to in data-space
# differs (fixed 0/0.5/1 vs each column's own min/median/max). So rather
# than a hand-picked "value > threshold" cutoff per scale type, work out
# the actual background colour Excel will render for a cell and choose
# text colour by its relative luminance - one rule, correct for both scales
# and immune to the two ever drifting apart again.
_SCALE_STOP_COLORS = ("FCFCFB", "86B6EF", "0D366B")


def _hex_to_rgb(h: str) -> tuple[float, float, float]:
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _lerp_rgb(t: float, c0: tuple, c1: tuple) -> tuple[float, float, float]:
    t = min(max(t, 0.0), 1.0)
    return tuple(c0[i] + (c1[i] - c0[i]) * t for i in range(3))


def _color_scale_rgb(value: float, vmin: float, vmid: float, vmax: float) -> tuple[float, float, float]:
    """Interpolated RGB for `value` under a 3-stop min/mid/max colour scale."""
    c0, c1, c2 = (_hex_to_rgb(c) for c in _SCALE_STOP_COLORS)
    if vmax <= vmin:
        return c0
    if value <= vmid:
        t = 0.0 if vmid <= vmin else (value - vmin) / (vmid - vmin)
        return _lerp_rgb(t, c0, c1)
    t = 0.0 if vmax <= vmid else (value - vmid) / (vmax - vmid)
    return _lerp_rgb(t, c1, c2)


def _relative_luminance(rgb: tuple[float, float, float]) -> float:
    r, g, b = (c / 255 for c in rgb)
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def _font_for_value(value: float, vmin: float, vmid: float, vmax: float) -> Font:
    """White text if the cell's own colour-scale background is dark enough
    to need it, black otherwise - computed from the actual interpolated
    colour rather than guessed from a raw-value threshold."""
    return WHITE_FONT if _relative_luminance(_color_scale_rgb(value, vmin, vmid, vmax)) < 0.5 else BLACK_FONT

# ---------------------------------------------------------------------------
# Legend content
# ---------------------------------------------------------------------------

COLUMN_DESCRIPTIONS = [
    ("country", "ISO3 country code (LUX = Luxembourg, DNK = Denmark)."),
    ("asset", "Infrastructure asset type: roads, airports, education, power."),
    ("scenario", "Modeling scenario - see the Scenarios table below."),
    ("outcome", "Which model output the row's indices describe - see the Outcomes table below."),
    ("factor", "The uncertainty factor being scored, in readable form. Curve factors (originally "
               "named e.g. curve_F7_4) are shown as a description instead - see the Curve_Groups "
               "sheet for the exact curve IDs and object types behind each one, and the "
               "factor_code column (where present) for the original internal name."),
    ("factor_code", "The original internal factor name (e.g. curve_F7_4, protection_scale) - use "
                     "this to cross-reference the Curve_Groups sheet or to filter/pivot programmatically."),
    ("S1", "First-order Sobol index: fraction of the outcome's variance explained by this factor "
           "acting ALONE (averaging out every other factor). Ranges roughly 0-1; small negative "
           "values are Monte Carlo noise around a true value near 0, not a real negative variance."),
    ("S1_conf", "95% confidence interval half-width on S1 (bootstrap, from SALib). If comparable "
                "in size to S1 itself, the estimate is not reliably distinguishable from 0 at this sample size."),
    ("ST", "Total-effect Sobol index: fraction of variance explained by this factor INCLUDING every "
           "interaction it has with other factors. Always >= S1."),
    ("ST_conf", "95% confidence interval half-width on ST."),
    ("interaction", "ST - S1: how much of this factor's influence only appears when combined with "
                     "other factors, rather than acting independently."),
    ("importance", "(All_Feature_Scores sheet only) Extra-trees feature importance from a Latin "
                    "Hypercube sample - a quicker, less rigorous alternative to the Sobol ST/S1 "
                    "columns. Useful as a sanity check, not a substitute."),
]

SCENARIO_DESCRIPTIONS = [
    ("(each scenario = ONE hazard)", "Hazards are never combined - every scenario computes exactly "
                                      "one of river flood / coastal flood / earthquake / windstorm, "
                                      "so their uncertainty structures are studied independently. The '_ds' "
                                      "suffix means depth error is MULTIPLICATIVE (depth_scale, "
                                      "x0.9-1.1) instead of the additive +/-0.5 m depth_offset - "
                                      "the two are otherwise identical, so a scenario and its _ds "
                                      "twin isolate the effect of how flood-depth error is modelled."),
    ("flood_baseline", "River flood only. Protection = FLOPROS design standard x protection_scale "
                        "(multiplier in [0, 2]). Additive depth error (depth_offset, +/-0.5 m)."),
    ("flood_baseline_ds", "As flood_baseline but depth error is multiplicative (depth_scale, x0.9-1.1)."),
    ("flood_absprot", "River flood only. Protection sampled as an ABSOLUTE return period "
                       "(protection_abs_rp, 5-200 years) applied uniformly to every feature, "
                       "replacing protection_scale - the only treatment that can move protection "
                       "for a feature whose FLOPROS baseline is 0 (0 x anything = 0). Additive depth."),
    ("flood_absprot_ds", "As flood_absprot but depth error is multiplicative (depth_scale, x0.9-1.1)."),
    ("flood_noprot", "River flood only. protection_scale held FIXED at 1.0 (exactly the FLOPROS "
                      "standards), not sampled - isolates how much the other flood factors matter "
                      "once protection uncertainty is set aside. Additive depth."),
    ("flood_noprot_ds", "As flood_noprot but depth error is multiplicative (depth_scale, x0.9-1.1)."),
    ("coastal_baseline / _absprot / _noprot (+ _ds twins)", "The six river-flood scenarios above, "
                     "mirrored for COASTAL flood: same three protection treatments x additive/"
                     "multiplicative depth, reusing the IDENTICAL flood depth-damage curves (so a "
                     "coastal scenario's curve factors still read 'Flood: ...'). Differences: the "
                     "coastal hazard maps (streamed from the CoCLiCo STAC catalogue), the COASTPROS "
                     "coastal protection standard instead of FLOPROS, and NO warming factor "
                     "(coastal sea-level rise is not modelled here). Coastal (non-landlocked) "
                     "countries only - absent for e.g. LUX."),
    ("earthquake", "Earthquake only. No protection standard (as in the reference). Factors: eq "
                    "curve choice(s), cost_level, pga_scale, aggregation."),
    ("windstorm", "Windstorm only. Fixed RP50 design-standard protection (IEC 60826), not sampled. "
                   "Factors: wind curve choice(s), cost_level, gust_scale, aggregation. Only defined "
                   "for wind-damaging assets - roads/ports carry no windstorm damage (their only wind "
                   "curve is identically zero) so the scenario is skipped for them."),
    ("windstorm_absprot", "As windstorm, but the design standard is sampled as an ABSOLUTE return "
                           "period (protection_abs_rp, 25-200 years) applied uniformly to every "
                           "feature, replacing the fixed RP50 - isolates the sensitivity to the "
                           "assumed wind design standard."),
]

ASSET_DESCRIPTIONS = [
    ("roads", "OSM road network (motorway down to track/service), line geometry."),
    ("airports", "Aerodrome/apron areas (polygon), runways (line), terminals (polygon)."),
    ("education", "Schools, kindergartens, colleges, universities, libraries (polygon building footprints)."),
    ("power", "Mixed geometry: lines/cables, point towers/poles/substations/transformers, "
              "polygon plants/generators/substations."),
    ("rail", "Rail and narrow-gauge track (line geometry)."),
    ("telecom", "Masts, towers, communications towers (point geometry)."),
    ("healthcare", "Hospitals and clinics (polygon building footprints)."),
    ("gas", "Gas pipelines (line), storage tanks and gasometers (polygon)."),
    ("oil", "Oil pipelines (line) and storage tanks (polygon)."),
    ("ports", "Ports and harbours (polygon). No windstorm (its only wind curve is zero)."),
]

FACTOR_DESCRIPTIONS = [
    ("warming", "Global warming level: current, 1.5C, 2.0C, 3.0C, 4.0C. Shifts every river flood "
                "return period via basin-level anchor maps (RP10/100/500 -> new RP)."),
    ("cost_level", "Reconstruction cost per unit, real-valued in [-1, 1]: -1 = the minimum cost "
                   "estimate, 0 = mean, +1 = maximum, piecewise-linear in between. Shared by "
                   "both hazards and every object type in the asset."),
    ("protection_scale", "Multiplier on the FLOPROS flood protection design standard, [0, 2]. "
                          "0 = no protection, 1 = FLOPROS estimate, 2 = double. baseline scenario only."),
    ("protection_abs_rp", "Absolute protection design standard in years, applied uniformly to every "
                           "feature regardless of its baseline. Flood/coastal '_absprot' scenarios "
                           "sample [5, 200] (replacing FLOPROS/COASTPROS); the 'windstorm_absprot' "
                           "scenario samples [25, 200] (replacing the fixed RP50)."),
    ("depth_offset", "Additive bias on river flood water depth, [-0.5, 0.5] metres. Can only shrink "
                      "or intensify damage within the already-mapped flood extent, not expand it. "
                      "Used by the non-'_ds' flood scenarios."),
    ("depth_scale", "Multiplicative factor on river flood water depth, [0.9, 1.1] (i.e. +/-10%). The "
                     "'_ds' flood scenarios use this INSTEAD of depth_offset, to test whether a "
                     "proportional depth error changes the sensitivity picture versus an additive one."),
    ("pga_scale", "Multiplier on earthquake ground motion (PGA), [0.8, 1.2] - a hazard-map "
                  "uncertainty factor, distinct from the fragility curve choice itself."),
    ("gust_scale", "Multiplier on windstorm 3-sec gust speed, [0.9, 1.1] - the windstorm hazard-map "
                    "uncertainty factor, analogous to pga_scale for earthquake."),
    ("aggregation", "Exposure aggregation order: 'per_cell' applies the damage curve to each "
                    "raster cell then sums (damagescanner's own approach); 'mean_depth' averages "
                    "intensity over the feature first, then applies the curve once."),
    ("Flood/Earthquake/Windstorm: ...", "A curve-choice factor. Shows which report classes or object "
                                      "types use this curve group, then in parentheses the "
                                      "group's lowest curve ID and how many curve options it "
                                      "offers - e.g. 'Flood: motorway_trunk/primary (F7.4, 4 "
                                      "curves)' means: applies to motorway/trunk/primary-class "
                                      "roads, and the sampled uncertainty is a choice among 4 "
                                      "flood curves starting at F7.4. A group with only 1 curve "
                                      "carries no uncertainty and was held fixed, not sampled - "
                                      "see the Curve_Groups sheet."),
]

TIMING_DESCRIPTIONS = [
    ("preprocess_s / validate_s", "Seconds spent in Stage 1 (GIS preprocessing) and validation "
                                   "for that (country, asset) - run once, shared by every scenario."),
    ("lhs_run_s / sobol_run_s", "Seconds spent actually running the LHS or Sobol experiments "
                                 "(ema_workbench evaluation, wall-clock, however many --workers "
                                 "were used at the time)."),
    ("lhs_analyze_s / sobol_analyze_s", "Seconds spent in the corresponding analyze.py/"
                                         "analyze_sobol.py step (feature scoring or SALib Sobol "
                                         "analysis + figures)."),
    ("lhs_n / sobol_n", "The sample size (--n) used for that run - matters when comparing timing "
                        "across reruns at different precision, e.g. an N=512 vs N=8192 Sobol run "
                        "for the same combination."),
    ("*_total_s", "Sum of the run+analyze steps it covers. Timing_By_Combo's combo_total_s adds "
                  "Stage 1 on top - the full time for that (country, asset) across every scenario."),
    ("Timing_Raw: n / ok / ts", "n = sample size used (where applicable); ok = whether the step "
                                 "succeeded; ts = UTC timestamp the step finished."),
    ("Source", "All three Timing_* sheets are parsed from results/run_study_log.jsonl, written by "
               "run_study.py. Only steps run through that orchestrator are logged - a standalone "
               "run_experiments.py call isn't. Where a step was rerun (e.g. a Sobol N bump), the "
               "By_Combo/Summary sheets show the MOST RECENT run only; Timing_Raw keeps every "
               "attempt."),
]

RERUN_DESCRIPTIONS = [
    ("How a rerun pair is found", "Every experiments_{country}_{asset}_{scenario}_sobol_n<N>_<timestamp>.tar.gz "
                                   "archive in results/ is grouped by (country, asset, scenario). A combo with "
                                   "archives at 2+ distinct N gets compared: smallest N = 'old', largest N = "
                                   "'new'. Combos only ever run once are not shown - nothing to compare yet. "
                                   "Always on total_EAD_MEUR, the headline outcome."),
    ("old_n / new_n", "The Sobol sample size N used for the earlier/later run (actual experiment count is "
                       "N x (2k+2) for k uncertain factors, per the Saltelli scheme)."),
    ("old_top_factor / new_top_factor", "(Reruns_Overview) Whichever factor had the single highest ST at "
                                         "that N. If this differs old-to-new, the extra runs didn't just "
                                         "tighten confidence, they changed which factor looks most important."),
    ("old_worst_ratio / new_worst_ratio", "(Reruns_Overview) The worst (largest) ST_conf/ST ratio among "
                                           "factors with ST > 0.05 at that N - one number summarising 'how "
                                           "trustworthy is the least-trustworthy relevant factor'."),
    ("top_factor_changed", "(Reruns_Overview) TRUE if old_top_factor != new_top_factor - the clearest "
                            "single sign that the original sample size wasn't enough to safely rank the "
                            "top drivers."),
    ("old_ratio / new_ratio", "(Reruns_Detail) ST_conf/ST for that specific factor at that N. Blank where "
                               "ST <= 0.05 (too small to matter, so precision on it isn't tracked). Lower "
                               "is more precise; both scales use the same fixed colouring so old vs new "
                               "colour intensity is directly comparable."),
    ("still_imprecise", "(Reruns_Detail) TRUE if, even at the higher new_n, ST > 0.05 and the ST_conf/ST "
                         "ratio is still above 0.3 - a candidate for yet another rerun at even higher N."),
]

CONVERGENCE_DESCRIPTIONS = [
    ("What Sobol_Convergence is", "One row per (country, asset, scenario) adaptive-Sobol search. "
                                   "Instead of a fixed sample size, the Sobol base N is doubled from "
                                   "--sobol-min-n (default 128) until the result is precise enough on "
                                   "total_EAD_MEUR, or --sobol-max-n (default 8192) is reached. Only "
                                   "combinations run through src.adaptive_sobol appear here."),
    ("stop_n / stop_reason", "The base sample size the search stopped at, and why: 'converged' (the "
                              "precision target below was met), 'max_n_reached' (hit the cap still "
                              "above target - treat its ranking as provisional), or "
                              "'no_relevant_factors' (no factor had ST>0.05, e.g. a hazard with no "
                              "exposure - nothing to resolve, stopped immediately)."),
    ("threshold / worst_ratio_at_stop", "The search stops once the worst ST_conf/ST among relevant "
                                          "factors (ST>0.05) drops below `threshold` (default 0.2). "
                                          "worst_ratio_at_stop is that worst ratio at the stopping N - "
                                          "if it's still above threshold, the run hit the N cap."),
    ("ratio_by_n", "The whole search in one cell: 'N:worst_ratio' for every N tried, e.g. "
                    "'128:0.41, 256:0.29, 512:0.18' means it converged at N=512. 'n/a' means no "
                    "relevant factor at that N."),
    ("n_factors / total_runs / elapsed_s", "n_factors = number of uncertainty factors (k). total_runs "
                                            "= model evaluations summed across every N tried "
                                            "(each N costs N x (2k+2) runs). elapsed_s = wall-clock for "
                                            "the whole adaptive search."),
]

OUTCOME_DESCRIPTIONS = [
    ("total_EAD_MEUR", "Total expected annual damage for this scenario's single hazard, million "
                        "EUR/year. (Scenarios are single-hazard, so this equals the one EAD_*_MEUR "
                        "hazard column below that the scenario computes.)"),
    ("EAD_river_MEUR", "Expected annual damage from river flooding (flood_* scenarios)."),
    ("EAD_coastal_MEUR", "Expected annual damage from coastal flooding (coastal_* scenarios)."),
    ("EAD_earthquake_MEUR", "Expected annual damage from earthquake (earthquake scenario)."),
    ("EAD_windstorm_MEUR", "Expected annual damage from windstorm (windstorm scenario)."),
    ("damage_RP100_river_MEUR", "Total damage from a single simulated 1-in-100-year flood event "
                                 "(not annualised - a snapshot, not an expectation)."),
    ("damage_RP100_coastal_MEUR", "As above for a 1-in-100-year coastal flood event."),
    ("damage_RP100_windstorm_MEUR", "As above for a 1-in-100-year windstorm event."),
    ("exposed_qty_RP100_river", "Total exposed quantity at the RP100 flood extent, in the "
                                 "asset's native units (metres/m^2/count depending on geometry mix)."),
    ("exposed_qty_RP100_coastal", "As above for the RP100 coastal flood extent."),
    ("exposed_qty_RP100_windstorm", "As above for the RP100 windstorm extent."),
    ("EAD_<class>_MEUR", "Expected annual damage broken down by report class: for roads, a "
                          "5-class road hierarchy (motorway_trunk/primary/secondary/tertiary/other); "
                          "for other assets, the raw OSM object_type."),
]


def _write_kv_table(ws: Worksheet, start_row: int, title: str, rows: list[tuple[str, str]],
                     col_widths: tuple[int, int] = (30, 100)) -> int:
    ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
    r = start_row + 1
    for key, desc in rows:
        ws.cell(row=r, column=1, value=key).font = Font(bold=True)
        c2 = ws.cell(row=r, column=2, value=desc)
        c2.alignment = WRAP
        r += 1
    ws.column_dimensions["A"].width = col_widths[0]
    ws.column_dimensions["B"].width = col_widths[1]
    return r + 1


def write_legend_sheet(ws: Worksheet) -> None:
    ws.cell(row=1, column=1, value="MIRACA uncertainty study - result glossary").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Every abbreviation used across the sheets in this workbook, explained.").font = Font(italic=True)
    r = 4
    r = _write_kv_table(ws, r, "Column meanings (All_Sobol_Indices / ST_* sheets)", COLUMN_DESCRIPTIONS)
    r = _write_kv_table(ws, r, "Scenarios", SCENARIO_DESCRIPTIONS)
    r = _write_kv_table(ws, r, "Asset types", ASSET_DESCRIPTIONS)
    r = _write_kv_table(ws, r, "Factors", FACTOR_DESCRIPTIONS)
    r = _write_kv_table(ws, r, "Outcomes", OUTCOME_DESCRIPTIONS)
    r = _write_kv_table(ws, r, "Sobol_Convergence columns", CONVERGENCE_DESCRIPTIONS)
    r = _write_kv_table(ws, r, "Reruns_Overview / Reruns_Detail columns", RERUN_DESCRIPTIONS)
    _write_kv_table(ws, r, "Timing_By_Combo / Timing_Summary / Timing_Raw columns", TIMING_DESCRIPTIONS)
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Curve group lookup + readable labels (built live from src.curves, never
# goes stale relative to the actual factor definitions)
# ---------------------------------------------------------------------------


def build_curve_groups_table() -> pd.DataFrame:
    rows = []
    for asset, cfg in ASSET_CONFIGS.items():
        for hazard, groups, obj_group in (
            ("river", cfg.flood_groups, cfg.flood_object_group),
            ("earthquake", cfg.eq_groups, cfg.eq_object_group),
            ("windstorm", cfg.wind_groups, cfg.wind_object_group),
        ):
            for group_name, curve_ids in sorted(groups.items()):
                members = sorted(o for o, g in obj_group.items() if g == group_name)
                samplable = len(curve_ids) > 1
                rows.append(
                    {
                        "asset": asset,
                        "hazard": hazard,
                        "factor_code": f"curve_{group_name}",
                        "readable_label": _curve_group_label(asset, hazard, group_name),
                        "curve_ids_in_group": ", ".join(curve_ids),
                        "n_curve_options": len(curve_ids),
                        "sampled_or_fixed": "sampled" if samplable else "FIXED (only 1 curve available)",
                        "object_types_using_this_group": ", ".join(members),
                    }
                )
    return pd.DataFrame(rows).sort_values(["asset", "hazard", "factor_code"])


def _short_member_list(members: list[str], max_items: int = 6) -> str:
    if len(members) <= max_items:
        return "/".join(members)
    return "/".join(members[:max_items]) + f" +{len(members) - max_items} more"


_HAZARD_GROUPS = {
    "river": ("flood_groups", "flood_object_group", "Flood"),
    "earthquake": ("eq_groups", "eq_object_group", "Earthquake"),
    "windstorm": ("wind_groups", "wind_object_group", "Windstorm"),
}


def _curve_group_label(asset: str, hazard: str, group_name: str) -> str:
    cfg = ASSET_CONFIGS[asset]
    groups_attr, obj_attr, hazard_word = _HAZARD_GROUPS[hazard]
    groups = getattr(cfg, groups_attr)
    obj_group = getattr(cfg, obj_attr)
    curve_ids = groups[group_name]
    members = sorted(o for o, g in obj_group.items() if g == group_name)

    if cfg.report_class is not None:
        classes = sorted({cfg.report_class.get(m, cfg.default_report_class) for m in members})
        desc = "all classes" if set(classes) == set(cfg.report_classes) else "/".join(classes)
    else:
        desc = _short_member_list(members)

    n = len(curve_ids)
    return f"{hazard_word}: {desc} ({curve_ids[0]}, {n} curve{'s' if n != 1 else ''})"


def _build_curve_label_lookup() -> dict[tuple[str, str], str]:
    lookup = {}
    for asset, cfg in ASSET_CONFIGS.items():
        for hazard, (groups_attr, _obj, _word) in _HAZARD_GROUPS.items():
            for group_name in getattr(cfg, groups_attr):
                lookup[(asset, f"curve_{group_name}")] = _curve_group_label(asset, hazard, group_name)
    return lookup


CURVE_LABELS = _build_curve_label_lookup()


def readable_factor(asset: str, factor_code: str) -> str:
    if factor_code.startswith("curve_"):
        return CURVE_LABELS.get((asset, factor_code), factor_code)
    return factor_code


# ---------------------------------------------------------------------------
# Result discovery + loading
# ---------------------------------------------------------------------------


def parse_result_filename(path: Path, suffix: str) -> tuple[str, str, str] | None:
    name = path.name
    if not name.endswith(suffix):
        return None
    stem = name[: -len(suffix)]
    if "_" not in stem:
        return None
    country, rest = stem.split("_", 1)
    for asset in ASSET_CONFIGS:
        prefix = f"{asset}_"
        if rest.startswith(prefix):
            scenario = rest[len(prefix):]
            if scenario in SCENARIOS:
                return country, asset, scenario
    return None


def load_all_sobol(results_dir: Path) -> pd.DataFrame:
    frames = []
    # rglob: per-combo CSVs now live in per-country subfolders (results/<ISO3>/),
    # but still pick up any legacy files written flat into results/.
    for path in sorted(results_dir.rglob("*_sobol_indices.csv")):
        parsed = parse_result_filename(path, "_sobol_indices.csv")
        if parsed is None:
            continue
        country, asset, scenario = parsed
        df = pd.read_csv(path)
        df.insert(0, "country", country)
        df.insert(1, "asset", asset)
        df.insert(2, "scenario", scenario)
        frames.append(df)
    cols = ["country", "asset", "scenario", "outcome", "factor", "S1", "S1_conf", "ST", "ST_conf", "interaction"]
    if not frames:
        return pd.DataFrame(columns=cols)
    out = pd.concat(frames, ignore_index=True)
    out = out.rename(columns={"factor": "factor_code"})
    out.insert(
        out.columns.get_loc("factor_code") + 1, "factor",
        [readable_factor(a, f) for a, f in zip(out["asset"], out["factor_code"])],
    )
    return out.sort_values(
        ["country", "asset", "scenario", "outcome", "ST"], ascending=[True, True, True, True, False]
    ).reset_index(drop=True)


def load_all_feature_scores(results_dir: Path) -> pd.DataFrame:
    frames = []
    for path in sorted(results_dir.rglob("*_feature_scores.csv")):
        parsed = parse_result_filename(path, "_feature_scores.csv")
        if parsed is None:
            continue
        country, asset, scenario = parsed
        df = pd.read_csv(path, index_col=0).rename_axis("factor_code").reset_index()
        long = df.melt(id_vars="factor_code", var_name="outcome", value_name="importance")
        long.insert(0, "country", country)
        long.insert(1, "asset", asset)
        long.insert(2, "scenario", scenario)
        frames.append(long)
    cols = ["country", "asset", "scenario", "factor_code", "factor", "outcome", "importance"]
    if not frames:
        return pd.DataFrame(columns=cols)
    out = pd.concat(frames, ignore_index=True)
    out.insert(
        out.columns.get_loc("factor_code") + 1, "factor",
        [readable_factor(a, f) for a, f in zip(out["asset"], out["factor_code"])],
    )
    return out.sort_values(
        ["country", "asset", "scenario", "outcome", "importance"], ascending=[True, True, True, True, False]
    ).reset_index(drop=True)


def build_top_drivers(sobol_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    total = sobol_df[sobol_df["outcome"] == "total_EAD_MEUR"]
    for (country, asset, scenario), g in total.groupby(["country", "asset", "scenario"]):
        g = g.sort_values("ST", ascending=False).reset_index(drop=True)

        def _at(i, col):
            return g.loc[i, col] if i < len(g) else (np.nan if col != "factor" else None)

        rows.append(
            {
                "asset": asset,
                "scenario": scenario,
                "country": country,
                "n_factors": len(g),
                "1st_factor": _at(0, "factor"), "1st_ST": _at(0, "ST"),
                "2nd_factor": _at(1, "factor"), "2nd_ST": _at(1, "ST"),
                "3rd_factor": _at(2, "factor"), "3rd_ST": _at(2, "ST"),
            }
        )
    asset_order = {a: i for i, a in enumerate(ASSET_CONFIGS)}
    scen_order = {s: i for i, s in enumerate(SCENARIOS)}
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["_a"] = df["asset"].map(asset_order)
    df["_s"] = df["scenario"].map(scen_order)
    return df.sort_values(["_a", "_s", "country"]).drop(columns=["_a", "_s"]).reset_index(drop=True)


def build_st_heatmap(sobol_df: pd.DataFrame, scenario: str) -> pd.DataFrame:
    sub = sobol_df[(sobol_df["scenario"] == scenario) & (sobol_df["outcome"] == "total_EAD_MEUR")].copy()
    if sub.empty:
        return pd.DataFrame()
    sub["combo"] = sub["asset"] + " / " + sub["country"]
    # Curve factors mean a different thing per asset (e.g. curve_F9_1 is
    # "aerodrome/apron" for airports but "terminal" for power), so heatmap
    # rows - which have no per-cell asset column to disambiguate - prefix
    # curve rows with the asset name. Generic factors (protection_scale,
    # cost_level, ...) mean the same thing everywhere and stay unprefixed,
    # so they compare directly across every asset's columns.
    sub["row_label"] = [
        f"[{a}] {readable_factor(a, code)}" if code.startswith("curve_") else code
        for a, code in zip(sub["asset"], sub["factor_code"])
    ]
    piv = sub.pivot_table(index="row_label", columns="combo", values="ST")
    asset_order = {a: i for i, a in enumerate(ASSET_CONFIGS)}
    combos = sorted(piv.columns, key=lambda c: (asset_order.get(c.split(" / ")[0], 99), c))
    # generic-factor rows first (readable across all assets), then curve rows grouped by asset
    rows = sorted(piv.index, key=lambda r: (r.startswith("["), r))
    return piv.reindex(index=rows, columns=combos)


# ---------------------------------------------------------------------------
# Timing (results/run_study_log.jsonl, written by run_study.py)
# ---------------------------------------------------------------------------

STAGE1_STEPS = ("preprocess", "validate")
# adaptive_sobol replaced the old fixed-N run_experiments_sobol step; both are
# accepted so timing sheets still parse older logs from before the change.
SCENARIO_STEPS = (
    "run_experiments_lhs", "analyze_lhs",
    "adaptive_sobol", "run_experiments_sobol", "analyze_sobol",
)
STEP_LABELS = {
    "preprocess": "preprocess_s",
    "validate": "validate_s",
    "run_experiments_lhs": "lhs_run_s",
    "analyze_lhs": "lhs_analyze_s",
    "adaptive_sobol": "sobol_run_s",
    "run_experiments_sobol": "sobol_run_s",
    "analyze_sobol": "sobol_analyze_s",
}


def load_timing_log(results_dir: Path) -> pd.DataFrame:
    """Parse results/run_study_log.jsonl into a tidy DataFrame.

    Tolerates a torn last line (the file may be being appended to by a study
    run still in progress when this is called) by skipping any line that
    fails to parse rather than raising.
    """
    path = results_dir / "run_study_log.jsonl"
    cols = ["country", "asset", "scenario", "step", "n", "ok", "elapsed_s", "ts"]
    if not path.exists():
        return pd.DataFrame(columns=cols)

    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            country, _, asset = rec.get("combo", "").partition("/")
            rows.append(
                {
                    "country": country,
                    "asset": asset,
                    "scenario": rec.get("scenario"),
                    "step": rec.get("step"),
                    "n": rec.get("n"),
                    "ok": rec.get("ok"),
                    "elapsed_s": rec.get("elapsed_s"),
                    "ts": rec.get("ts"),
                }
            )
    df = pd.DataFrame(rows, columns=cols)
    return df.sort_values("ts").reset_index(drop=True) if len(df) else df


def build_timing_by_combo(timing_df: pd.DataFrame) -> pd.DataFrame:
    """One row per (country, asset): Stage-1 time + summed scenario time."""
    if timing_df.empty:
        return pd.DataFrame()

    stage1 = timing_df[timing_df["step"].isin(STAGE1_STEPS)]
    stage1 = stage1.sort_values("ts").drop_duplicates(["country", "asset", "step"], keep="last")
    stage1_piv = stage1.pivot_table(
        index=["country", "asset"], columns="step", values="elapsed_s", aggfunc="last"
    ).rename(columns=STEP_LABELS)

    scen = timing_df[timing_df["step"].isin(SCENARIO_STEPS)]
    scen = scen.sort_values("ts").drop_duplicates(["country", "asset", "scenario", "step"], keep="last")
    scen_total = scen.groupby(["country", "asset"])["elapsed_s"].sum().rename("all_scenarios_s")

    out = pd.concat([stage1_piv, scen_total], axis=1).reset_index()
    for col in ("preprocess_s", "validate_s", "all_scenarios_s"):
        if col not in out.columns:
            out[col] = np.nan
    out["combo_total_s"] = out[["preprocess_s", "validate_s", "all_scenarios_s"]].sum(axis=1, skipna=True)

    asset_order = {a: i for i, a in enumerate(ASSET_CONFIGS)}
    out["_a"] = out["asset"].map(asset_order)
    out = out.sort_values(["_a", "country"]).drop(columns="_a").reset_index(drop=True)
    return out[["country", "asset", "preprocess_s", "validate_s", "all_scenarios_s", "combo_total_s"]]


def build_timing_summary(timing_df: pd.DataFrame) -> pd.DataFrame:
    """One row per (country, asset, scenario): time per LHS/Sobol run+analyze step."""
    if timing_df.empty:
        return pd.DataFrame()

    scen = timing_df[timing_df["step"].isin(SCENARIO_STEPS)].copy()
    scen = scen.sort_values("ts").drop_duplicates(["country", "asset", "scenario", "step"], keep="last")

    piv = scen.pivot_table(
        index=["country", "asset", "scenario"], columns="step", values="elapsed_s", aggfunc="last"
    ).rename(columns=STEP_LABELS).reset_index()
    for col in STEP_LABELS.values():
        if col not in piv.columns and col not in ("preprocess_s", "validate_s"):
            piv[col] = np.nan
    run_cols = ["lhs_run_s", "lhs_analyze_s", "sobol_run_s", "sobol_analyze_s"]
    piv["scenario_total_s"] = piv[run_cols].sum(axis=1, skipna=True)

    n_key = ["country", "asset", "scenario"]
    lhs_n = scen[scen["step"] == "run_experiments_lhs"].set_index(n_key)["n"]
    sobol_n = scen[scen["step"] == "run_experiments_sobol"].set_index(n_key)["n"]
    idx = pd.MultiIndex.from_frame(piv[n_key])
    piv["lhs_n"] = idx.map(lhs_n)
    piv["sobol_n"] = idx.map(sobol_n)

    asset_order = {a: i for i, a in enumerate(ASSET_CONFIGS)}
    scen_order = {s: i for i, s in enumerate(SCENARIOS)}
    piv["_a"] = piv["asset"].map(asset_order)
    piv["_s"] = piv["scenario"].map(scen_order)
    piv = piv.sort_values(["_a", "_s", "country"]).drop(columns=["_a", "_s"]).reset_index(drop=True)

    cols = ["country", "asset", "scenario", "lhs_n", "lhs_run_s", "lhs_analyze_s",
            "sobol_n", "sobol_run_s", "sobol_analyze_s", "scenario_total_s"]
    return piv[[c for c in cols if c in piv.columns]]


# ---------------------------------------------------------------------------
# Adaptive-Sobol convergence (results/sobol_convergence_log.jsonl)
# ---------------------------------------------------------------------------


def load_convergence_log(results_dir: Path) -> pd.DataFrame:
    """One row per (country, asset, scenario) adaptive-Sobol search.

    Reads results/sobol_convergence_log.jsonl (written by src.adaptive_sobol),
    keeping the most recent record per combination. Flattens the per-round
    detail into a compact 'ratio_by_n' string plus the headline stop info, so
    a limited test run can be read at a glance: where each combination's
    sample size stopped, and why.
    """
    path = results_dir / "sobol_convergence_log.jsonl"
    cols = ["country", "asset", "scenario", "stop_n", "stop_reason", "n_factors",
            "threshold", "worst_ratio_at_stop", "top_factor_at_stop", "n_rounds",
            "ratio_by_n", "total_runs", "elapsed_s", "ts"]
    if not path.exists():
        return pd.DataFrame(columns=cols)

    records = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            rounds = rec.get("rounds", []) or []

            def _fmt(r):
                wr = r.get("worst_ratio")
                return f"{r.get('n')}:{'n/a' if wr is None else round(wr, 3)}"

            last = rounds[-1] if rounds else {}
            records.append(
                {
                    "country": rec.get("country"),
                    "asset": rec.get("asset"),
                    "scenario": rec.get("scenario"),
                    "stop_n": rec.get("stop_n"),
                    "stop_reason": rec.get("stop_reason"),
                    "n_factors": rec.get("n_factors"),
                    "threshold": rec.get("threshold"),
                    "worst_ratio_at_stop": last.get("worst_ratio"),
                    "top_factor_at_stop": last.get("top_factor"),
                    "n_rounds": len(rounds),
                    "ratio_by_n": ", ".join(_fmt(r) for r in rounds),
                    "total_runs": sum(int(r.get("n_runs", 0) or 0) for r in rounds),
                    "elapsed_s": rec.get("elapsed_s"),
                    "ts": rec.get("ts"),
                }
            )
    df = pd.DataFrame(records, columns=cols)
    if df.empty:
        return df
    df = df.sort_values("ts").drop_duplicates(["country", "asset", "scenario"], keep="last")

    asset_order = {a: i for i, a in enumerate(ASSET_CONFIGS)}
    scen_order = {s: i for i, s in enumerate(SCENARIOS)}
    df["_a"] = df["asset"].map(asset_order)
    df["_s"] = df["scenario"].map(scen_order)
    df["factor"] = [readable_factor(a, str(f)) if f else f
                    for a, f in zip(df["asset"], df["top_factor_at_stop"])]
    df["top_factor_at_stop"] = df["factor"]
    return (
        df.sort_values(["_a", "_s", "country"])
        .drop(columns=["_a", "_s", "factor"])
        .reset_index(drop=True)
    )


# ---------------------------------------------------------------------------
# Sobol N-rerun comparison (before/after precision, e.g. N=512 -> N=8192)
# ---------------------------------------------------------------------------

RERUN_RELEVANCE = 0.05  # ST above this is "big enough to matter" for ranking/flagging
RERUN_TARGET_RATIO = 0.3  # ST_conf/ST above this still counts as imprecise

ARCHIVE_RE = re.compile(
    r"^experiments_(?P<country>[A-Z]{3})_(?P<rest>.+)_sobol_n(?P<n>\d+)_\d{8}_\d{6}\.tar\.gz$"
)


def find_sobol_rerun_pairs(results_dir: Path) -> list[dict]:
    """Group every Sobol experiment archive by (country, asset, scenario).

    For any combination with archives at more than one sample size N, return
    the smallest-N ("old") and largest-N ("new") archive - the two points a
    before/after precision comparison is built from. Combinations only ever
    run once (a single N) are skipped - there is nothing to compare yet.
    Fully general: picks up whatever reruns exist for any country/asset/
    scenario, not just the LUX-only batch that motivated this feature.
    """
    groups: dict[tuple[str, str, str], dict[int, Path]] = {}
    for path in results_dir.rglob("experiments_*_sobol_n*_*.tar.gz"):
        m = ARCHIVE_RE.match(path.name)
        if not m:
            continue
        country, rest, n = m.group("country"), m.group("rest"), int(m.group("n"))
        asset = scenario = None
        for a in ASSET_CONFIGS:
            prefix = f"{a}_"
            if rest.startswith(prefix):
                asset, scenario = a, rest[len(prefix):]
                break
        if asset is None or scenario not in SCENARIOS:
            continue
        by_n = groups.setdefault((country, asset, scenario), {})
        existing = by_n.get(n)
        if existing is None or path.stat().st_mtime > existing.stat().st_mtime:
            by_n[n] = path

    pairs = []
    for (country, asset, scenario), by_n in groups.items():
        if len(by_n) < 2:
            continue
        ns = sorted(by_n)
        pairs.append(
            {
                "country": country, "asset": asset, "scenario": scenario,
                "old_n": ns[0], "old_path": by_n[ns[0]],
                "new_n": ns[-1], "new_path": by_n[ns[-1]],
            }
        )
    return pairs


def _analyze_archive(path: Path, problem: dict, outcome: str) -> pd.DataFrame | None:
    """Re-run SALib's Sobol analysis on an archived experiment set.

    Returns None if the outcome has ~zero variance (e.g. an asset/hazard
    combination with no exposure at all) - Sobol indices are meaningless
    there, same edge case already handled elsewhere in this module.
    """
    from ema_workbench import load_results

    _, outcomes = load_results(str(path))
    y = np.asarray(outcomes[outcome], dtype=float)
    if np.var(y) < 1e-30:
        return None
    from SALib.analyze import sobol as salib_sobol

    si = salib_sobol.analyze(problem, y, calc_second_order=True, print_to_console=False)
    return pd.DataFrame(
        {
            "factor_code": problem["names"],
            "S1": np.nan_to_num(si["S1"]),
            "ST": np.nan_to_num(si["ST"]),
            "ST_conf": np.nan_to_num(si["ST_conf"]),
        }
    )


def build_rerun_comparisons(
    results_dir: Path, outcome: str = "total_EAD_MEUR"
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Before/after Sobol precision comparison for every rerun pair found by
    find_sobol_rerun_pairs(), on `outcome` only (total_EAD_MEUR - the
    headline result used throughout the rest of this workbook).

    Returns (overview, detail):
      overview - one row per rerun combo: old/new sample size, old/new #1
                  driver and its ST, and the worst ST_conf/ST ratio among
                  relevant (ST > RERUN_RELEVANCE) factors at each N - the
                  "did more runs actually change the headline answer" view.
      detail   - one row per rerun combo x factor (every factor, not just
                  the top one): old/new S1/ST/ST_conf and the precision
                  ratio, plus whether it's still imprecise after the rerun.
    """
    pairs = find_sobol_rerun_pairs(results_dir)
    overview_rows, detail_rows = [], []

    for p in pairs:
        set_country_override(p["country"])
        set_asset_override(p["asset"])
        set_scenario_override(p["scenario"])
        cfg = load_config()
        model = build_model(cfg)
        from ema_workbench.em_framework.salib_samplers import get_SALib_problem

        problem = get_SALib_problem(model.uncertainties)

        old_df = _analyze_archive(p["old_path"], problem, outcome)
        new_df = _analyze_archive(p["new_path"], problem, outcome)

        base = {"country": p["country"], "asset": p["asset"], "scenario": p["scenario"]}
        if old_df is None or new_df is None:
            # Zero-variance outcome at this combo (e.g. no exposure at all) -
            # nothing meaningful to compare, but still record that it was checked.
            overview_rows.append(
                {
                    **base,
                    "old_n": p["old_n"], "old_top_factor": None, "old_top_ST": 0.0, "old_worst_ratio": 0.0,
                    "new_n": p["new_n"], "new_top_factor": None, "new_top_ST": 0.0, "new_worst_ratio": 0.0,
                    "top_factor_changed": False,
                }
            )
            continue

        merged = old_df.merge(new_df, on="factor_code", suffixes=("_old", "_new"))
        merged["ratio_old"] = merged["ST_conf_old"] / merged["ST_old"].abs()
        merged["ratio_new"] = merged["ST_conf_new"] / merged["ST_new"].abs()
        merged["factor"] = [readable_factor(p["asset"], c) for c in merged["factor_code"]]

        for _, row in merged.iterrows():
            detail_rows.append(
                {
                    **base,
                    "factor_code": row["factor_code"], "factor": row["factor"],
                    "old_n": p["old_n"], "old_S1": row["S1_old"], "old_ST": row["ST_old"],
                    "old_ST_conf": row["ST_conf_old"],
                    "old_ratio": row["ratio_old"] if row["ST_old"] > RERUN_RELEVANCE else np.nan,
                    "new_n": p["new_n"], "new_S1": row["S1_new"], "new_ST": row["ST_new"],
                    "new_ST_conf": row["ST_conf_new"],
                    "new_ratio": row["ratio_new"] if row["ST_new"] > RERUN_RELEVANCE else np.nan,
                    "still_imprecise": bool(
                        row["ST_new"] > RERUN_RELEVANCE and row["ratio_new"] > RERUN_TARGET_RATIO
                    ),
                }
            )

        relevant_old = merged[merged["ST_old"] > RERUN_RELEVANCE]
        relevant_new = merged[merged["ST_new"] > RERUN_RELEVANCE]
        old_top = merged.loc[merged["ST_old"].idxmax()]
        new_top = merged.loc[merged["ST_new"].idxmax()]
        overview_rows.append(
            {
                **base,
                "old_n": p["old_n"], "old_top_factor": old_top["factor"], "old_top_ST": old_top["ST_old"],
                "old_worst_ratio": float(relevant_old["ratio_old"].max()) if len(relevant_old) else 0.0,
                "new_n": p["new_n"], "new_top_factor": new_top["factor"], "new_top_ST": new_top["ST_new"],
                "new_worst_ratio": float(relevant_new["ratio_new"].max()) if len(relevant_new) else 0.0,
                "top_factor_changed": bool(old_top["factor_code"] != new_top["factor_code"]),
            }
        )

    asset_order = {a: i for i, a in enumerate(ASSET_CONFIGS)}
    scen_order = {s: i for i, s in enumerate(SCENARIOS)}
    overview = pd.DataFrame(overview_rows)
    detail = pd.DataFrame(detail_rows)
    for d, extra_sort in ((overview, []), (detail, ["factor"])):
        if d.empty:
            continue
        d["_a"] = d["asset"].map(asset_order)
        d["_s"] = d["scenario"].map(scen_order)
        d.sort_values(["_a", "_s", "country"] + extra_sort, inplace=True)
        d.drop(columns=["_a", "_s"], inplace=True)
        d.reset_index(drop=True, inplace=True)
    return overview, detail


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------


def _apply_fraction_formatting(ws: Worksheet, rng: str) -> None:
    """Fixed 0/0.5/1 colour scale + white text on cells dark enough to need it.

    Reads the font decision straight back from each cell's own persisted
    value (rather than a separately-passed parallel array) so it can never
    drift out of alignment with what Excel actually displays.
    """
    for row_cells in ws[rng]:
        for cell in row_cells:
            cell.number_format = "0.000"
            val = cell.value
            if val is None or val == "":
                continue
            cell.font = _font_for_value(float(val), 0.0, 0.5, 1.0)
    ws.conditional_formatting.add(rng, FRACTION_COLOR_SCALE)


def _apply_relative_formatting(ws: Worksheet, rng: str, values) -> None:
    """Percentile-based colour scale (min/median/max of the column's own
    values, matching RELATIVE_COLOR_SCALE) + white text on cells dark
    enough to need it, using the same luminance rule as the fixed scale.

    `values` must be in the same row order the cells in `rng` were written
    in (i.e. the DataFrame column that produced them).
    """
    finite = pd.to_numeric(pd.Series(values), errors="coerce").dropna()
    vmin = float(finite.min()) if len(finite) else 0.0
    vmax = float(finite.max()) if len(finite) else 1.0
    vmid = float(finite.median()) if len(finite) else 0.5
    cells = [c for row in ws[rng] for c in row]
    for cell, val in zip(cells, values):
        cell.number_format = "0.000"
        if val is None or (isinstance(val, float) and np.isnan(val)):
            continue
        cell.font = _font_for_value(float(val), vmin, vmid, vmax)
    ws.conditional_formatting.add(rng, RELATIVE_COLOR_SCALE)


def _style_data_sheet(
    ws: Worksheet, df: pd.DataFrame,
    fraction_cols: list[str] | None = None,
    relative_cols: list[str] | None = None,
) -> None:
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        max_val_len = max((len(str(v)) for v in df[col]), default=10) if len(df) else 10
        width = max(10, min(42, max_val_len + 2), len(str(col)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in fraction_cols or []:
        if col not in df.columns:
            continue
        col_idx = list(df.columns).index(col) + 1
        letter = get_column_letter(col_idx)
        rng = f"{letter}2:{letter}{len(df) + 1}"
        _apply_fraction_formatting(ws, rng)

    for col in relative_cols or []:
        if col not in df.columns:
            continue
        col_idx = list(df.columns).index(col) + 1
        letter = get_column_letter(col_idx)
        rng = f"{letter}2:{letter}{len(df) + 1}"
        _apply_relative_formatting(ws, rng, df[col].tolist())


def _write_heatmap_sheet(writer: pd.ExcelWriter, sheet_name: str, piv: pd.DataFrame, scenario: str) -> None:
    if piv.empty:
        return
    piv.round(4).to_excel(writer, sheet_name=sheet_name, startrow=2)
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1,
            value=f"Sobol total-effect (ST) on total_EAD_MEUR - scenario: {scenario} "
                  f"(thick border = that column's single most decisive factor)").font = TITLE_FONT
    header_row = 3
    ws.cell(row=header_row, column=1).font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    for j in range(len(piv.columns)):
        c = ws.cell(row=header_row, column=2 + j)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws.column_dimensions["A"].width = 46
    for j in range(len(piv.columns)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 15

    n_rows, n_cols = piv.shape
    if n_rows and n_cols:
        rng = f"B{header_row + 1}:{get_column_letter(1 + n_cols)}{header_row + n_rows}"
        _apply_fraction_formatting(ws, rng)

        # Border around each column's single highest-ST cell - skipped when
        # every value in the column is (near) zero, e.g. a hazard with no
        # exposure at all (LUX airports has zero river-flood exposure, so
        # flood_noprot_ds's ST column is all zeros): argmax would pick
        # an arbitrary row with no real meaning, so mark nothing instead.
        for j, col in enumerate(piv.columns):
            series = piv[col].dropna()
            if series.empty or series.max() <= 1e-9:
                continue
            row_pos = list(piv.index).index(series.idxmax())
            ws.cell(row=header_row + 1 + row_pos, column=2 + j).border = TOP_FACTOR_BORDER

    ws.freeze_panes = f"B{header_row + 1}"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=str, default=None,
                        help=f"output .xlsx path (default: {PROJECT_ROOT / OUTPUT_NAME})")
    args = parser.parse_args()

    cfg = load_config()
    results_dir = cfg["results_dir"]
    out_path = Path(args.output) if args.output else PROJECT_ROOT / OUTPUT_NAME

    print(f"Scanning {results_dir} ...")
    sobol_df = load_all_sobol(results_dir)
    fscore_df = load_all_feature_scores(results_dir)
    n_combos = sobol_df[["country", "asset", "scenario"]].drop_duplicates().shape[0] if len(sobol_df) else 0
    print(f"  {len(sobol_df)} Sobol index rows across {n_combos} combinations")
    print(f"  {len(fscore_df)} feature-score rows")

    if sobol_df.empty:
        raise SystemExit(f"No *_sobol_indices.csv found in {results_dir} - run the study first.")

    top_drivers = build_top_drivers(sobol_df)
    curve_groups = build_curve_groups_table()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        writer.book.create_sheet("Legend")
        write_legend_sheet(writer.book["Legend"])

        curve_groups.to_excel(writer, sheet_name="Curve_Groups", index=False)
        _style_data_sheet(writer.sheets["Curve_Groups"], curve_groups)

        top_drivers.to_excel(writer, sheet_name="Top_Drivers", index=False)
        _style_data_sheet(
            writer.sheets["Top_Drivers"], top_drivers,
            fraction_cols=["1st_ST", "2nd_ST", "3rd_ST"],
        )

        for scenario in SCENARIOS:
            sheet_name = f"ST_{scenario}"[:31]  # Excel sheet name limit
            piv = build_st_heatmap(sobol_df, scenario)
            _write_heatmap_sheet(writer, sheet_name, piv, scenario)

        sobol_df.round(6).to_excel(writer, sheet_name="All_Sobol_Indices", index=False)
        _style_data_sheet(
            writer.sheets["All_Sobol_Indices"], sobol_df,
            fraction_cols=["S1", "ST"],
            relative_cols=["S1_conf", "ST_conf", "interaction"],
        )

        if not fscore_df.empty:
            fscore_df.round(6).to_excel(writer, sheet_name="All_Feature_Scores", index=False)
            _style_data_sheet(
                writer.sheets["All_Feature_Scores"], fscore_df, fraction_cols=["importance"]
            )

        convergence_df = load_convergence_log(results_dir)
        convergence_sheets = []
        if not convergence_df.empty:
            print(f"  {len(convergence_df)} adaptive-Sobol convergence records")
            convergence_df.to_excel(writer, sheet_name="Sobol_Convergence", index=False)
            _style_data_sheet(
                writer.sheets["Sobol_Convergence"], convergence_df,
                fraction_cols=["threshold", "worst_ratio_at_stop"],
                relative_cols=["stop_n", "total_runs", "elapsed_s"],
            )
            convergence_sheets.append("Sobol_Convergence")
        else:
            print("  no results/sobol_convergence_log.jsonl - skipping Sobol_Convergence sheet")

        print("  checking for Sobol reruns at different sample sizes...")
        rerun_overview, rerun_detail = build_rerun_comparisons(results_dir)
        rerun_sheets = []
        if not rerun_overview.empty:
            print(f"  {len(rerun_overview)} rerun combo(s) found, {len(rerun_detail)} factor rows")
            rerun_overview.round(4).to_excel(writer, sheet_name="Reruns_Overview", index=False)
            _style_data_sheet(
                writer.sheets["Reruns_Overview"], rerun_overview,
                fraction_cols=["old_top_ST", "old_worst_ratio", "new_top_ST", "new_worst_ratio"],
            )
            rerun_sheets.append("Reruns_Overview")

            rerun_detail.round(4).to_excel(writer, sheet_name="Reruns_Detail", index=False)
            _style_data_sheet(
                writer.sheets["Reruns_Detail"], rerun_detail,
                fraction_cols=["old_S1", "old_ST", "old_ratio", "new_S1", "new_ST", "new_ratio"],
                relative_cols=["old_ST_conf", "new_ST_conf"],
            )
            rerun_sheets.append("Reruns_Detail")
        else:
            print("  no combination has more than one Sobol N yet - skipping Reruns_* sheets")

        timing_df = load_timing_log(results_dir)
        print(f"  {len(timing_df)} timing log rows")
        timing_sheets = []
        if not timing_df.empty:
            by_combo = build_timing_by_combo(timing_df)
            summary = build_timing_summary(timing_df)
            elapsed_cols_combo = ["preprocess_s", "validate_s", "all_scenarios_s", "combo_total_s"]
            elapsed_cols_summary = ["lhs_run_s", "lhs_analyze_s", "sobol_run_s", "sobol_analyze_s", "scenario_total_s"]

            by_combo.round(1).to_excel(writer, sheet_name="Timing_By_Combo", index=False)
            _style_data_sheet(writer.sheets["Timing_By_Combo"], by_combo, relative_cols=elapsed_cols_combo)
            timing_sheets.append("Timing_By_Combo")

            summary.round(1).to_excel(writer, sheet_name="Timing_Summary", index=False)
            _style_data_sheet(writer.sheets["Timing_Summary"], summary, relative_cols=elapsed_cols_summary)
            timing_sheets.append("Timing_Summary")

            timing_df.to_excel(writer, sheet_name="Timing_Raw", index=False)
            _style_data_sheet(writer.sheets["Timing_Raw"], timing_df, relative_cols=["elapsed_s"])
            timing_sheets.append("Timing_Raw")
        else:
            print("  no results/run_study_log.jsonl found - skipping Timing sheets "
                  "(only produced when run_study.py has been used)")

        order = (
            ["Legend", "Curve_Groups", "Top_Drivers"]
            + [f"ST_{s}"[:31] for s in SCENARIOS]
            + ["All_Sobol_Indices", "All_Feature_Scores"]
            + convergence_sheets
            + rerun_sheets
            + timing_sheets
        )
        writer.book._sheets = [writer.book[name] for name in order if name in writer.book.sheetnames]

    print(f"\nSaved {out_path}")


if __name__ == "__main__":
    main()
